#encoding=utf-8
import openpyxl
import time
from config.VarConfig import *
class ParseExcel(object):
    def __init__(self):
        self.wb=None
        self.excelFile=None
    def loadWorkBook(self,excelPath):
        self.wb=openpyxl.load_workbook(excelPath)
        self.excelFile=excelPath
        return self.wb
    def getSheetByName(self,sheetName):
        sheet=self.wb.get_sheet_by_name(sheetName)
        return sheet
    def getSheetByIndex(self,sheetIndex):
        sheetName=self.wb.get_sheet_names()[sheetIndex]
        sheet=self.wb.get_sheet_by_name(sheetName)
        return sheet
    def getCellOfValue(self,sheet,rowNo,colsNo):
        return sheet.cell(row=rowNo,column=colsNo).value
    def writeCell(self,sheet,content,rowNo,colsNo):
        sheet.cell(row=rowNo, column=colsNo).value=content
        self.wb.save(self.excelFile)
    def getCurrentTime(self):
        return time.strftime("%Y-%m-%d %H:%M:%S",time.localtime(int(time.time())))
    def getRowsNum(self,sheet):
        return sheet.max_row
if __name__=='__main__':

    pe=ParseExcel()
    pe.loadWorkBook(excelPath)
    sheet=pe.getSheetByIndex(0)
    print(pe.getCellOfValue(sheet,rowNo=1,colsNo=1))
    sheet=pe.getSheetByName(sheetName)
    print(pe.getCellOfValue(sheet, rowNo=2, colsNo=7))
    pe.writeCell(sheet,'没有', rowNo=2, colsNo=7)
    print(pe.getCellOfValue(sheet, rowNo=2, colsNo=7))
    print(pe.getCurrentTime())
    print(sheet.max_row)
    print(pe.getRowsNum(sheet))